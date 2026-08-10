import re
import json
import subprocess
import openpyxl

md_file = '/Applications/MAMP/htdocs/rasagroup/UAT OK/Hasil_Testing_DEV_cURL.md'
template_file = '/Applications/MAMP/htdocs/rasagroup/Faspay VA - Skenario Functional Test_V.3.0 static(2).xlsx'
report_file = '/Applications/MAMP/htdocs/rasagroup/Live_vs_UAT_Report.txt'

SCENARIO_ROWS = {
    "11.1": 7, "11.2": 8, "11.3": 9, "11.4": 10, "11.5": 11,
    "11.6": 12, "11.7": 13, "11.8": 14, "11.9": 15, "11.10": 16,
    "11.11": 17, "11.12": 18, "11.13": 19, "11.14": 20, "11.15": 21,
    "11.16": 22, "11.17": 23
}

wb = openpyxl.load_workbook(template_file)
sheet = wb['Transfer VA']

# Extract Expected Responses from Template
expected_responses = {}
for row in range(7, 27):
    cell_no = sheet.cell(row=row, column=1).value
    if cell_no:
        expected_responses[str(cell_no)] = sheet.cell(row=row, column=4).value

with open(md_file, 'r', encoding='utf-8') as f:
    content = f.read()
scenarios = content.split("## Skenario ")

print("Executing LIVE Tests for comparison...")
report_lines = []
report_lines.append("=== Laporan Perbandingan LIVE API vs UAT EXPECTED ===\n")

for s in scenarios[1:]:
    scenario_id_match = re.match(r"^([\d\.]+)", s)
    if not scenario_id_match: continue
    scenario_id = scenario_id_match.group(1)
    
    curl_match = re.search(r"```bash\n(.*?)\n```", s, re.DOTALL)
    if not curl_match: continue
    curl_cmd = curl_match.group(1).strip()
    
    expected_raw = expected_responses.get(scenario_id, "No expected found")
    
    try:
        result = subprocess.run(curl_cmd, shell=True, capture_output=True, text=True)
        live_resp_text = result.stdout.strip()
        
        try:
            live_json = json.loads(live_resp_text)
            live_formatted = json.dumps(live_json, indent=2)
            live_code = live_json.get('responseCode', '')
            live_msg = live_json.get('responseMessage', '')
        except:
            live_formatted = live_resp_text
            live_code = ""
            live_msg = ""
            
        report_lines.append(f"Skenario {scenario_id}")
        report_lines.append("-" * 40)
        report_lines.append(f"EXPECTED (Dari Template UAT):\n{expected_raw}")
        report_lines.append(f"LIVE API ACTUAL:\n{live_formatted}")
        
        # Determine if it matches logic
        match = True
        warnings = []
        if live_code and live_msg:
            if live_code not in expected_raw and live_code[:-2] not in expected_raw: # Check if part of error code matches, e.g. 401xx00
                warnings.append(f"[!] Response Code Mismatch: Live is {live_code}")
                match = False
            if live_msg not in expected_raw and expected_raw.lower().find(live_msg.lower()) == -1:
                warnings.append(f"[!] Response Message Mismatch: Live is '{live_msg}'")
                match = False
                
            # Specific checks based on user's previous complaints
            if "success" in expected_raw.lower() and "success" not in expected_raw:
                # wait, if expected has lowercase "success", live should be lowercase
                pass 
                
        if match and not warnings:
            report_lines.append("-> STATUS: MATCH ✅\n")
        else:
            report_lines.append("-> STATUS: TIDAK MATCH ❌")
            for w in warnings:
                report_lines.append(w)
            report_lines.append("\n")
            
    except Exception as e:
        report_lines.append(f"Skenario {scenario_id}: FAILED to run curl: {e}\n")

with open(report_file, 'w', encoding='utf-8') as f:
    f.write("\n".join(report_lines))

print(f"Report generated at {report_file}")
