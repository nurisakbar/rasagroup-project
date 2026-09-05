import openpyxl
import json

wb = openpyxl.load_workbook('/Applications/MAMP/htdocs/rasagroup/faspay_uat_1108_dev.xlsx', data_only=True)
sheet = wb['Transfer VA']

scenarios_to_check = ['11.2', '11.17']
for row in range(7, 30):
    scenario = sheet.cell(row=row, column=1).value
    if scenario and str(scenario).strip() in scenarios_to_check:
        print("Scenario:", scenario)
        print("--- REQUEST ---")
        print(sheet.cell(row=row, column=5).value)
        print("--- RESPONSE ---")
        print(sheet.cell(row=row, column=6).value)
        print("="*40)
