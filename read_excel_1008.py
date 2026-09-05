import openpyxl

wb = openpyxl.load_workbook('/Applications/MAMP/htdocs/rasagroup/faspay_uat_1008sore.xlsx', data_only=True)
sheet = wb['Transfer VA']

for row in range(7, 30):
    scenario = sheet.cell(row=row, column=1).value
    if scenario and str(scenario).strip() == '11.17':
        print("Scenario:", scenario)
        print("--- REQUEST ---")
        print(sheet.cell(row=row, column=5).value)
        print("--- RESPONSE ---")
        print(sheet.cell(row=row, column=6).value)
        break
