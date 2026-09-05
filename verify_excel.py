import openpyxl

wb = openpyxl.load_workbook('/Applications/MAMP/htdocs/rasagroup/faspay_uat_1108_dev.xlsx', data_only=True)
sheet = wb['Transfer VA']

print("=== VERIFICATION OF faspay_uat_1108_dev.xlsx ===")
for row in range(7, 30):
    scenario = str(sheet.cell(row=row, column=1).value).strip()
    if scenario in ['11.1', '11.2', '11.10', '11.17']:
        print(f"\n--- SCENARIO {scenario} ---")
        print("REQUEST CELL:")
        print(sheet.cell(row=row, column=5).value)
        print("\nRESPONSE CELL:")
        print(sheet.cell(row=row, column=6).value)
        print("=" * 50)
