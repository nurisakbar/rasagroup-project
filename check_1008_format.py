import openpyxl

wb = openpyxl.load_workbook('/Applications/MAMP/htdocs/rasagroup/faspay_uat_1008sore.xlsx', data_only=True)
sheet = wb['Transfer VA']

for row in range(6, 12):
    scenario = sheet.cell(row=row, column=1).value
    print("Row:", row, "Scenario:", scenario)
    print("REQUEST:")
    print(sheet.cell(row=row, column=5).value)
    print("\nRESPONSE:")
    print(sheet.cell(row=row, column=6).value)
    print("-" * 40)
