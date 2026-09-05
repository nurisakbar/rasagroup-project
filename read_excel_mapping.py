import openpyxl

file_path = '/Applications/MAMP/htdocs/rasagroup/Mapping Menu Admin Ecommerce - PIC RASA Connect.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets:", wb.sheetnames)
    sheet = wb.active
    
    print("\n--- Content from Sheet:", sheet.title, "---")
    
    # Try to find headers
    for row in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
        if any(row):
            print(row)
            
except Exception as e:
    print(f"Error: {e}")
