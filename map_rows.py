import openpyxl
wb = openpyxl.load_workbook('/Applications/MAMP/htdocs/rasagroup/Faspay VA - Skenario Functional Test_V.3.0 static(2).xlsx')
sheet = wb['Transfer VA']
for r in range(6, 26):
    print(f"Row {r}: A='{sheet.cell(r, 1).value}', B='{sheet.cell(r, 2).value}', C='{sheet.cell(r, 3).value}'")
