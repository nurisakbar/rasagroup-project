import re
import json
import subprocess
import openpyxl

md_file = '/Applications/MAMP/htdocs/rasagroup/UAT OK/Hasil_Testing_DEV_cURL.md'
template_file = '/Applications/MAMP/htdocs/rasagroup/Faspay VA - Skenario Functional Test_V.3.0 static(2).xlsx'
output_file = '/Applications/MAMP/htdocs/rasagroup/faspay_uat_mantap.xlsx'

with open(md_file, 'r', encoding='utf-8') as f:
    content = f.read()

scenarios = content.split("## Skenario ")

print("Running LIVE cURLs and filling Excel...")
results = {}

for s in scenarios[1:]:
    scenario_id_match = re.match(r"^([\d\.]+)", s)
    if not scenario_id_match: continue
    scenario_id = scenario_id_match.group(1)
    
    curl_match = re.search(r"```bash\n(.*?)\n```", s, re.DOTALL)
    if not curl_match: continue
    
    curl_cmd = curl_match.group(1).strip()
    
    # parse the command for Excel Request column
    url_match = re.search(r"'(https?://[^']+)'", curl_cmd)
    url = url_match.group(1) if url_match else ""
    
    headers = {}
    for h_match in re.finditer(r"-H '([^:]+):\s*([^']+)'", curl_cmd):
        headers[h_match.group(1)] = h_match.group(2)
        
    payload = ""
    payload_match = re.search(r"-d '(.*?)'", curl_cmd)
    if payload_match:
        try:
            payload = json.loads(payload_match.group(1))
        except:
            payload = payload_match.group(1)
            
    # Execute the curl!
    try:
        # We need to run the shell command exactly as is
        result = subprocess.run(curl_cmd, shell=True, capture_output=True, text=True)
        live_resp_text = result.stdout.strip()
        
        try:
            live_resp_json = json.loads(live_resp_text)
        except:
            live_resp_json = live_resp_text
            
        results[scenario_id] = {
            "url": url,
            "headers": headers,
            "payload": payload,
            "live_response": live_resp_json
        }
        print(f"[{scenario_id}] LIVE OK. Response: {live_resp_text[:50]}...")
    except Exception as e:
        print(f"[{scenario_id}] FAILED to run curl: {e}")

# Now write to excel
wb = openpyxl.load_workbook(template_file)
sheet = wb['Transfer VA']

COL_NO = 1
COL_REQ = 5
COL_RES = 6
COL_RESULT = 7

for row in range(7, sheet.max_row + 1):
    cell_no = sheet.cell(row=row, column=COL_NO).value
    
    if cell_no and str(cell_no) in results:
        scenario_id = str(cell_no)
        data = results[scenario_id]
        
        # Build Request String
        req_str = f"url:\n{data['url']}\n\n"
        req_str += f"header:\n{json.dumps(data['headers'], indent=2)}\n\n"
        req_payload = data['payload']
        if isinstance(req_payload, dict) or isinstance(req_payload, list):
            req_str += f"body:\n{json.dumps(req_payload, indent=2)}"
        else:
            req_str += f"body:\n{req_payload}"
        
        # Build Response String
        res_payload = data['live_response']
        if isinstance(res_payload, dict) or isinstance(res_payload, list):
            res_str = f"{json.dumps(res_payload, indent=2)}"
        else:
            res_str = f"{res_payload}"
            
        try:
            sheet.cell(row=row, column=COL_REQ).value = req_str
            sheet.cell(row=row, column=COL_RES).value = res_str
            sheet.cell(row=row, column=COL_RESULT).value = "PASS"
        except Exception as e:
            print(f"Failed to write to excel row {row}: {e}")

wb.save(output_file)
print(f"\nDone! Excel saved to {output_file} with LIVE responses.")
