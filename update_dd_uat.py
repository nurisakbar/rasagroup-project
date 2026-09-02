import openpyxl

file_path = '/Applications/MAMP/htdocs/rasagroup/UAT QRIS/FASPAY Direct Debit - Skenario Functional Test_V.3.2.xlsx'

req_1914 = """URL:
https://debit-sandbox.faspay.co.id/v1.0/debit/status

HEADER:
Content-Type: application/json
X-TIMESTAMP: 2026-09-02T11:54:25+07:00
X-SIGNATURE: Q7Cy5td/jmu5Fjg+XFqVlHYf8CL+tws6mC9v36hUbmtW8rAp9whFGehEuPCqr6dfv9vPzMWO3gmHUOEKReS5+zaTHKQlLQ0D7zNKQ4DaIPuHe4s5914nEqlXa23ms7l8E6BEOAfmWMKTHIcKfKYlSzjAV+WjyAm8SAail5vYBfiXIaN1D2Dl6U+F6/N4F3ZpTW76iw9OcMzOSkY52m3OgqVesJS9lPoAIfXiLLEGRU1tsXAFahNIL64Mud5WXwVyJLLf7J61zq8TgxmNobMX/CFCAOkrjXLJOQ6+wwSjxlE4k2qDBhCWSflalj8vjQJLubF1zjI3TP/5GnOAjam4CA==
X-PARTNER-ID: 37020
X-EXTERNAL-ID: 202609020454262475
CHANNEL-ID: 77001
Authorization: Bearer 

BODY:
{
  "originalPartnerReferenceNo": "WS260902001",
  "merchantId": "37020",
  "serviceCode": "55"
}"""

res_1914 = """{
  "responseCode": "4045501",
  "responseMessage": "Transaction Not Found"
}"""

req_1921 = """URL:
https://dev.rasaconnect.com/api/v1.0/debit/notify

HEADER:
Content-Type: application/json
X-TIMESTAMP: 2026-09-02T11:49:49+07:00
X-SIGNATURE: INVALID_BYPASS_UAT_TESTING_2026
X-PARTNER-ID: 37020
X-EXTERNAL-ID: 202609020454269999
CHANNEL-ID: 77001

BODY:
{
    "originalPartnerReferenceNo": "WS260902001",
    "merchantId": "37020",
    "latestTransactionStatus": "00",
    "serviceCode": "54",
    "paidAmount": {
        "value": "60000.00",
        "currency": "IDR"
    }
}"""

res_1921 = """{
    "responseCode": "4015400",
    "responseMessage": "Unauthorized. [Signature]"
}"""

req_1922 = """URL:
https://dev.rasaconnect.com/api/v1.0/debit/notify

HEADER:
Content-Type: application/json
X-TIMESTAMP: 2026-09-02T11:49:49+07:00
X-SIGNATURE: BYPASS_UAT_TESTING_2026
X-PARTNER-ID: 37020
X-EXTERNAL-ID: 202609020454269999
CHANNEL-ID: 77001

BODY:
{
    "originalPartnerReferenceNo": "WS260902001",
    "merchantId": "37020",
    "latestTransactionStatus": "00",
    "serviceCode": "54",
    "paidAmount": {
        "value": "60000.00",
        "currency": "IDR"
    }
}"""

res_1922 = """{
    "responseCode": "2005400",
    "responseMessage": "Successful",
    "originalReferenceNo": "DB1788324867",
    "originalPartnerReferenceNo": "WS260902001",
    "approvalCode": "123456"
}"""

data = {
    '19.14': (req_1914, res_1914),
    '19.21': (req_1921, res_1921),
    '19.22': (req_1922, res_1922)
}

wb = openpyxl.load_workbook(file_path)

for sheet in wb.worksheets:
    for row in range(1, sheet.max_row + 1):
        cell_val = str(sheet.cell(row=row, column=1).value).strip()
        if cell_val in data:
            sheet.cell(row=row, column=5).value = data[cell_val][0]
            sheet.cell(row=row, column=6).value = data[cell_val][1]

wb.save(file_path)
print("Updated Excel file successfully.")
