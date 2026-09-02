import openpyxl

file_path = '/Applications/MAMP/htdocs/rasagroup/UAT QRIS/FASPAY Direct Debit - Skenario Functional Test_V.3.2.xlsx'
wb = openpyxl.load_workbook(file_path)

for sheet in wb.worksheets:
    for row in range(1, sheet.max_row + 1):
        cell_val = str(sheet.cell(row=row, column=1).value).strip()
        
        if cell_val == '19.3':
            sheet.cell(row=row, column=4).value = "4005402\nBad Request. Missing Mandatory Field [merchantId]"
        elif cell_val == '19.4':
            sheet.cell(row=row, column=4).value = "4005401\nBad Request. Invalid Field Format [amount.value]"
        elif cell_val == '19.14':
            sheet.cell(row=row, column=4).value = "4045501\nTransaction Not Found"
            
            # Update Request body to use WS260902999
            req = str(sheet.cell(row=row, column=5).value)
            req = req.replace("WS260902001", "WS260902999")
            sheet.cell(row=row, column=5).value = req

wb.save(file_path)
print("Refinements applied successfully.")
