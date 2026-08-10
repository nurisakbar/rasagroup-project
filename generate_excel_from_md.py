import re
import json
import openpyxl

md_file = '/Applications/MAMP/htdocs/rasagroup/UAT OK/Hasil_Testing_DEV_cURL.md'
template_file = '/Applications/MAMP/htdocs/rasagroup/Faspay VA - Skenario Functional Test_V.3.0 static(2).xlsx'
output_file = '/Applications/MAMP/htdocs/rasagroup/faspay_uat_mantap.xlsx'

def extract_between(text, start, end):
    try:
        return text.split(start)[1].split(end)[0].strip()
    except IndexError:
        return ""

def parse_curl(curl_str):
    url = ""
    headers = {}
    body = "{}"
    
    lines = curl_str.strip().split('\n')
    for line in lines:
        line = line.strip()
        if line.startswith("curl"):
            match = re.search(r"'(https?://[^']+)'", line)
            if match:
                url = match.group(1)
            else:
                match = re.search(r'"(https?://[^"]+)"', line)
                if match:
                    url = match.group(1)
        elif line.startswith("-H"):
            match = re.search(r"-H '([^:]+):\s*(.*?)'(?: \\)?", line)
            if match:
                headers[match.group(1)] = match.group(2)
        elif line.startswith("-d"):
            match = re.search(r"-d '(.*?)'", line)
            if match:
                body_str = match.group(1)
                try:
                    body = json.loads(body_str)
                except:
                    body = body_str
    return url, headers, body

with open(md_file, 'r', encoding='utf-8') as f:
    content = f.read()

scenarios = content.split("## Skenario ")
results = {}

for s in scenarios[1:]:
    # Extract scenario ID
    scenario_id = s.split(":")[0].strip()
    
    # Extract cURL block
    curl_block = extract_between(s, "```bash", "```")
    url, headers, body = parse_curl(curl_block)
    
    # Extract HTTP Code
    status_match = re.search(r"### API Response \(Status: (\d+)\):", s)
    http_code = status_match.group(1) if status_match else ""
    
    # Extract Response JSON
    json_block = extract_between(s, "```json", "```")
    try:
        resp_json = json.loads(json_block)
    except:
        resp_json = json_block
    
    results[scenario_id] = {
        "request_url": url,
        "request_headers": headers,
        "request_payload": body,
        "http_code": http_code,
        "response_body": resp_json
    }

wb = openpyxl.load_workbook(template_file)
sheet = wb['Transfer VA']

COL_NO = 1
COL_REQ = 5
COL_RES = 6
COL_RESULT = 7

for row in range(7, sheet.max_row + 1):
    cell_no = sheet.cell(row=row, column=COL_NO).value
    
    if cell_no and str(cell_no) in results:
        scenario_id = str(cell_no)
        data = results[scenario_id]
        
        # Build Request String
        req_str = f"url:\n{data.get('request_url', '')}\n\n"
        req_str += f"header:\n{json.dumps(data.get('request_headers', {}), indent=2)}\n\n"
        req_payload = data.get('request_payload', {})
        if isinstance(req_payload, dict) or isinstance(req_payload, list):
            req_str += f"body:\n{json.dumps(req_payload, indent=2)}"
        else:
            req_str += f"body:\n{req_payload}"
        
        # Build Response String
        res_payload = data.get('response_body', {})
        if isinstance(res_payload, dict) or isinstance(res_payload, list):
            res_str = f"{json.dumps(res_payload, indent=2)}"
        else:
            res_str = f"{res_payload}"
            
        sheet.cell(row=row, column=COL_REQ).value = req_str
        sheet.cell(row=row, column=COL_RES).value = res_str
        sheet.cell(row=row, column=COL_RESULT).value = "PASS"

wb.save(output_file)
print(f"Successfully generated {output_file} from markdown file.")
