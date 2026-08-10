import openpyxl

template_file = '/Applications/MAMP/htdocs/rasagroup/Faspay VA - Skenario Functional Test_V.3.0 static(2).xlsx'
wb = openpyxl.load_workbook(template_file)
sheet = wb['Transfer VA']

print("Expected Responses from Template:")
for row in range(7, 25):
    cell_no = sheet.cell(row=row, column=1).value
    expected = sheet.cell(row=row, column=4).value
    if cell_no and expected:
        print(f"[{cell_no}] {expected}")
