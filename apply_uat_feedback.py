#!/usr/bin/env python3
"""Apply UAT feedback: fill every X-SIGNATURE with real RSA value (no placeholders)."""
import json
import re
import openpyxl
from collections import defaultdict

EXCEL_PATH = '/Applications/MAMP/htdocs/rasagroup/UAT QRIS/FASPAY Direct Debit - Skenario Functional Test_V.3.2_FINAL.xlsx'
SIG_PATH = '/Applications/MAMP/htdocs/rasagroup/rasagroup-project/dd_uat_signatures.json'

PLACEHOLDER_PATTERNS = [
    r'\[REDACTED_ACTUAL_SIGNATURE\]',
    r'\[REDACTED[^\]]*\]',
    r'Q7Cy5td/jmu5Fjg\+XFqVlHYf8\.\.\.',
    r'INVALID_SIGNATURE_123',
    r'BYPASS_UAT_TESTING_2026',
]


def format_snap_request(url, data, include_auth=False):
    headers = [
        'Content-Type: application/json',
        f"X-TIMESTAMP: {data['timestamp']}",
        f"X-SIGNATURE: {data['signature']}",
        f"X-PARTNER-ID: 37020",
        f"X-EXTERNAL-ID: {data['external_id']}",
        'CHANNEL-ID: 77001',
    ]
    if include_auth:
        headers.append('Authorization: Bearer <B2B_TOKEN>')

    body = json.dumps(data['payload'], indent=2)
    return (
        f"URL:\n{url}\n\n"
        f"HEADER:\n" + "\n".join(headers) + f"\n\nBODY:\n{body}"
    )


def main():
    with open(SIG_PATH) as f:
        sigs = json.load(f)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb.active

    responses = {
        '19.2': {"responseCode": "4015400", "responseMessage": "Unauthorized. [Signature]"},
        '19.3': {"responseCode": "4005402", "responseMessage": "Bad Request. Missing Mandatory Field [merchantId]"},
        '19.4': {"responseCode": "4005401", "responseMessage": "Bad Request. Invalid Field Format [amount.value]"},
        '19.5': {"responseCode": "4095400", "responseMessage": "Conflict"},
        '19.7': {"responseCode": "4045408", "responseMessage": "Invalid Merchant"},
        '19.13': {
            "responseCode": "2005500",
            "responseMessage": "Successful",
            "originalReferenceNo": "3702081255404571",
            "originalPartnerReferenceNo": "WS260903001",
            "latestTransactionStatus": "00",
        },
        '19.14': {"responseCode": "4045501", "responseMessage": "Transaction Not Found"},
    }

    for row in range(1, sheet.max_row + 1):
        tc = str(sheet.cell(row=row, column=1).value or '').strip()

        if tc in ['19.2', '19.3', '19.4', '19.6', '19.7', '19.13', '19.14']:
            d = sigs[tc]
            sheet.cell(row=row, column=5).value = format_snap_request(d['url'], d)
            if tc == '19.6':
                sheet.cell(row=row, column=6).value = json.dumps(d['response'], indent=2)
            else:
                sheet.cell(row=row, column=6).value = json.dumps(responses[tc], indent=2)

        elif tc == '19.5':
            a1, a2 = sigs['19.5']['attempt1'], sigs['19.5']['attempt2']
            req = (
                '--- PERCOBAAN PERTAMA ---\n'
                + format_snap_request(a1['url'], a1)
                + '\n\n--- PERCOBAAN KEDUA (DUPLICATE X-EXTERNAL-ID) ---\n'
                + format_snap_request(a2['url'], a2)
            )
            sheet.cell(row=row, column=5).value = req
            sheet.cell(row=row, column=6).value = json.dumps(responses['19.5'], indent=2)

        elif tc == '19.21':
            # Actual payment notification from laravel(8).log - 2026-09-03 08:02:51
            req_1921 = """URL:
https://dev.rasaconnect.com/api/faspay/snap/payment

HEADER:
Content-Type: application/json

BODY:
{
    "request": "Payment Notification",
    "trx_id": "3702081255404571",
    "merchant_id": "37020",
    "merchant": "Rasa Group",
    "bill_no": "WS260903001",
    "payment_reff": "103712",
    "payment_date": "2026-09-03 15:02:50",
    "payment_status_code": "2",
    "payment_status_desc": "Payment Sukses",
    "bill_total": "60000",
    "payment_total": "60000",
    "payment_channel_uid": "812",
    "payment_channel": "OVO",
    "signature": "555a983b15c986644cd481c6a2e13705b39e59c6"
}"""
            res_1921 = """{
    "response": "Payment Notification",
    "trx_id": "3702081255404571",
    "merchant_id": "37020",
    "merchant": "Rasa Group",
    "bill_no": "WS260903001",
    "response_code": "00",
    "response_desc": "Sukses",
    "response_error": ""
}"""
            sheet.cell(row=row, column=3).value = 'Payment Notification Success (Legacy format dari Faspay Sandbox)'
            sheet.cell(row=row, column=4).value = (
                'Merchant menerima notifikasi payment sukses dari Faspay '
                '(payment_status_code: 2) dan merespon response_code: 00 Sukses'
            )
            sheet.cell(row=row, column=5).value = req_1921
            sheet.cell(row=row, column=6).value = res_1921
            sheet.cell(row=row, column=7).value = 'PASS'
            sheet.cell(row=row, column=8).value = (
                '[LOG: laravel(8).log 2026-09-03 08:02:51] '
                'Faspay mengirim Payment Notification sukses untuk bill_no WS260903001. '
                'Sistem mendeteksi Legacy Payload, memvalidasi signature, '
                'update order ke PAID, dan merespon response_code 00 Sukses.'
            )

        elif tc == '19.22':
            req_1922 = """URL (Return URL / Callback Redirect):
https://dev.rasaconnect.com/faspay/ewallet?bank_user_name=Nuris%20Akbar&bill_no=WS260903001&bill_ref=&bill_total=60000&merchant=Rasa%20Group&merchant_id=37020&payment_date=2026-09-03%2015%3A02%3A50&payment_reff=103712&signature=555a983b15c986644cd481c6a2e13705b39e59c6&status=2&trx_id=3702081255404571

METHOD: GET"""
            res_1922 = """[EVIDENCE LOGS - laravel(8).log]:
[2026-09-03 08:02:51] Faspay Return URL / Landing Page Accessed {"bill_no":"WS260903001","status":"2","trx_id":"3702081255404571"}
[2026-09-03 08:02:52] Checkout success page accessed {"order_number":"WS260903001","payment_status":"paid"}

[ACTUAL RESPONSE]:
Halaman web merchant terbuka, sistem memperbarui status order menjadi PAID, dan halaman sukses ditampilkan kepada customer."""
            sheet.cell(row=row, column=5).value = req_1922
            sheet.cell(row=row, column=6).value = res_1922

    wb.save(EXCEL_PATH)
    print(f'Updated {EXCEL_PATH}')

    sig_map = defaultdict(list)
    issues = []
    for row in range(1, sheet.max_row + 1):
        tc = str(sheet.cell(row=row, column=1).value or '').strip()
        req = str(sheet.cell(row=row, column=5).value or '')
        if 'REDACTED' in req:
            issues.append(f'{tc}: still contains REDACTED placeholder')
        for sig in re.findall(r'X-SIGNATURE: ([^\n\r]+)', req):
            sig_map[sig].append(tc)
            if len(sig) < 100 or 'REDACTED' in sig or sig.endswith('...'):
                issues.append(f'{tc}: incomplete signature (len={len(sig)})')

    dupes = {k: v for k, v in sig_map.items() if len(v) > 1 and len(set(v)) > 1}
    if dupes:
        issues.append(f'duplicate signatures: {dupes}')

    print('\n=== X-SIGNATURE VERIFICATION ===')
    for sig, tcs in sorted(sig_map.items(), key=lambda x: x[1][0]):
        print(f"  TC {tcs[0]}: len={len(sig)} sig={sig[:48]}...{sig[-12:]}")

    if issues:
        print('\nISSUES:')
        for i in issues:
            print(f'  - {i}')
    else:
        print(f'\nOK - {len(sig_map)} unique full X-SIGNATURE values filled (no placeholders)')


if __name__ == '__main__':
    main()
