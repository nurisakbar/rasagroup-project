import openpyxl
import json

wb = openpyxl.load_workbook('/Applications/MAMP/htdocs/rasagroup/faspay_uat_1108_dev.xlsx', data_only=True)
sheet = wb['Transfer VA']

print("=== EXPECTED VS ACTUAL RESPONSE COMPARISON ===\n")

for row in range(10, 50): # Assuming test cases are within rows 10-50
    scenario_id = sheet.cell(row=row, column=1).value
    if not scenario_id or not str(scenario_id).startswith('11.'):
        continue
        
    expected_result = sheet.cell(row=row, column=4).value
    actual_response = sheet.cell(row=row, column=6).value
    
    print(f"--- SCENARIO {scenario_id} ---")
    print(f"EXPECTED (Col 4):\n{expected_result}")
    print(f"ACTUAL (Col 6):\n{actual_response}")
    
    # Try to decode ACTUAL to see if it matches EXPECTED's format
    try:
        if actual_response:
            resp_json = json.loads(actual_response)
            code = resp_json.get('responseCode', '')
            msg = resp_json.get('responseMessage', '')
            print(f"Parsed Actual -> Code: {code}, Msg: {msg}")
    except:
        pass
        
    print("="*50 + "\n")
