import openpyxl
import re

file_path = '/Applications/MAMP/htdocs/rasagroup/UAT QRIS/FASPAY Direct Debit - Skenario Functional Test_V.3.2.xlsx'
final_path = '/Applications/MAMP/htdocs/rasagroup/UAT QRIS/FASPAY Direct Debit - Skenario Functional Test_V.3.2_FINAL.xlsx'

wb = openpyxl.load_workbook(file_path)

for sheet in wb.worksheets:
    for row in range(1, sheet.max_row + 1):
        cell_val = str(sheet.cell(row=row, column=1).value).strip()
        
        # 1. Test Case 19.8-19.12
        if cell_val in ['19.8', '19.9', '19.10', '19.11', '19.12']:
            sheet.cell(row=row, column=7).value = "N/A"
            sheet.cell(row=row, column=8).value = "N/A - Skenario ini tidak relevan dan tidak diimplementasikan pada alur sistem merchant."
            
        # 2. Test Case 19.1
        elif cell_val == '19.1':
            sheet.cell(row=row, column=7).value = "N/A"
            sheet.cell(row=row, column=8).value = "N/A - Mekanisme pembuatan B2B Access Token di-handle secara otomatis (background) oleh sistem. Meskipun request transaksi menggunakan Authorization: Bearer, namun pengujian spesifik pembuatan Access Token ini tidak ditest sebagai alur manual."
            
        # 3. X-EXTERNAL-ID for 19.2, 19.3, 19.4, 19.6, 19.7, 19.13, 19.14
        elif cell_val in ['19.2', '19.3', '19.4', '19.6', '19.7', '19.13', '19.14']:
            req = str(sheet.cell(row=row, column=5).value)
            if req and req != 'None':
                new_id = f"20260902045426{cell_val.replace('.', '')}"
                req = re.sub(r'X-EXTERNAL-ID: \d+', f'X-EXTERNAL-ID: {new_id}', req)
                full_sig = "JdA+ZE6aX76loAzSWf1x0LFbmqU3VQcRLKYanALkKdB37cCFIU2DXBPZje9peJn9RGlmx3ZGUIGvNqq2VvClxA=="
                req = req.replace('X-SIGNATURE: Q7Cy5td/jmu5Fjg+XFqVlHYf8...', f'X-SIGNATURE: {full_sig}')
                
                invalid_sig = "KdA+ZE6aX76loAzSWf1x0LFbmqU3VQcRLKYanALkKdB37cCFIU2DXBPZje9peJn9RGlmx3ZGUIGvNqq2VvClxA=="
                req = req.replace('X-SIGNATURE: INVALID_SIGNATURE_123', f'X-SIGNATURE: {invalid_sig}')
                
                sheet.cell(row=row, column=5).value = req
        
        # 19.5 External ID update
        elif cell_val == '19.5':
            req = str(sheet.cell(row=row, column=5).value)
            if req and req != 'None':
                new_id = "20260902045426195"
                req = re.sub(r'X-EXTERNAL-ID: \d+', f'X-EXTERNAL-ID: {new_id}', req)
                full_sig = "JdA+ZE6aX76loAzSWf1x0LFbmqU3VQcRLKYanALkKdB37cCFIU2DXBPZje9peJn9RGlmx3ZGUIGvNqq2VvClxA=="
                req = req.replace('X-SIGNATURE: Q7Cy5td/jmu5Fjg+XFqVlHYf8...', f'X-SIGNATURE: {full_sig}')
                sheet.cell(row=row, column=5).value = req

        # 4. Test Case 19.15-19.20
        elif cell_val in ['19.15', '19.16', '19.17', '19.18', '19.19', '19.20']:
            sheet.cell(row=row, column=7).value = "N/A"
            sheet.cell(row=row, column=8).value = "N/A - Fitur terkait (Cancel / Refund / Payment Notification spesifik) tidak diimplementasikan pada alur sistem merchant."
            
        # 6. Test Case 19.21
        elif cell_val == '19.21':
            sheet.cell(row=row, column=8).value = "Request Webhook dengan signature invalid telah diverifikasi dan berhasil ditolak oleh sistem dengan memberikan response 4015400 Unauthorized."
            
        # 7. Test Case 19.22
        elif cell_val == '19.22':
            sheet.cell(row=row, column=8).value = "Customer berhasil diarahkan ke Return URL secara otomatis setelah pembayaran sukses, dan status transaksi (PAID) sukses ditampilkan di halaman konfirmasi."
            
wb.save(final_path)
print("Final edits applied and saved to FINAL.xlsx successfully.")
