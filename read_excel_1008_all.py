import openpyxl
import json

wb = openpyxl.load_workbook('/Applications/MAMP/htdocs/rasagroup/faspay_uat_1008sore.xlsx', data_only=True)
sheet = wb['Transfer VA']

scenarios = {}
for row in range(7, 30):
    scenario = sheet.cell(row=row, column=1).value
    if scenario:
        scenarios[str(scenario).strip()] = {
            'req': sheet.cell(row=row, column=5).value,
            'res': sheet.cell(row=row, column=6).value
        }

with open('extracted_1008.json', 'w') as f:
    json.dump(scenarios, f, indent=2)
