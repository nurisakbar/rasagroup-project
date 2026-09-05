import openpyxl

wb = openpyxl.load_workbook('/Applications/MAMP/htdocs/rasagroup/faspay_uat_1108_dev.xlsx', data_only=True)
sheet = wb['Transfer VA']

print("Scanning for errors...")
for row in range(7, 30):
    scenario = str(sheet.cell(row=row, column=1).value).strip()
    res = sheet.cell(row=row, column=6).value
    
    if res and ("Exception" in str(res) or "Error" in str(res) or "message" in str(res) and "responseCode" not in str(res)):
        print(f"--- SCENARIO {scenario} ---")
        print("RESPONSE CELL:")
        print(res)
        print("=" * 50)
