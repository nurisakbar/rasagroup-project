import openpyxl
from openpyxl.styles import Alignment
import json

excel_path = '/Users/nurisakbar/Downloads/Faspay VA - Skenario Functional Test_V.3.0 static(2).xlsx'
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

with open('uat.json', 'r') as f:
    data = json.load(f)

for row in range(9, sheet.max_row + 1):
    cell_no = sheet.cell(row=row, column=1).value
    if cell_no:
        tc_id = str(cell_no).strip()
        if tc_id in data['results']:
            tc = data['results'][tc_id]
            
            # format request string
            curl = f"url:\n{tc['request_url']}\n\nheader:\n"
            for k, v in tc['request_headers'].items():
                curl += f"{k}: {v}\n"
            curl += f"\nbody:\n{json.dumps(tc['request_payload'])}"
            
            # write values
            sheet.cell(row=row, column=5).value = curl
            sheet.cell(row=row, column=6).value = json.dumps(tc['response_body'], indent=2)
            sheet.cell(row=row, column=7).value = 'Pass'
            sheet.cell(row=row, column=8).value = ''
            
            # format cells
            align = Alignment(wrap_text=True, vertical='top')
            sheet.cell(row=row, column=5).alignment = align
            sheet.cell(row=row, column=6).alignment = align
            sheet.cell(row=row, column=7).alignment = Alignment(vertical='top')

wb.save(excel_path)
print("Excel file successfully updated!")
